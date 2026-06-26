"""
导出服务 — CSV、Excel、PDF 导出
对应需求文档 4.2.9
"""
import csv
import io
from typing import List, Optional
from datetime import date
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """数据导出服务"""

    @staticmethod
    def export_posts_csv(posts: List[dict]) -> io.StringIO:
        """导出帖子为 CSV 格式"""
        output = io.StringIO()
        # UTF-8 BOM for Excel compatibility
        output.write("﻿")

        fieldnames = [
            "id", "title", "category", "priority", "sentiment",
            "view_count", "reply_count", "author_name", "source",
            "tags", "is_anomaly", "risk_level", "data_date", "created_at"
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        # 中文表头
        header_map = {
            "id": "ID",
            "title": "标题",
            "category": "分类",
            "priority": "优先级",
            "sentiment": "情绪",
            "view_count": "浏览量",
            "reply_count": "回复数",
            "author_name": "作者",
            "source": "来源",
            "tags": "标签",
            "is_anomaly": "是否异常",
            "risk_level": "风险等级",
            "data_date": "数据日期",
            "created_at": "创建时间",
        }
        writer.writerow(header_map)

        for post in posts:
            row = {k: post.get(k, "") for k in fieldnames}
            # 将列表转为逗号分隔字符串
            for key in ["tags"]:
                val = row.get(key, [])
                if isinstance(val, list):
                    row[key] = ", ".join(str(v) for v in val)
            writer.writerow(row)

        output.seek(0)
        return output

    @staticmethod
    def export_posts_excel(posts: List[dict]) -> io.BytesIO:
        """导出帖子为 Excel 格式"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "帖子数据"

        # 样式定义
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell_alignment = Alignment(vertical="center")

        # 表头
        headers = [
            ("ID", 8), ("标题", 50), ("分类", 18), ("优先级", 10),
            ("情绪", 10), ("浏览量", 10), ("回复数", 10), ("作者", 15),
            ("来源", 15), ("是否异常", 10), ("风险等级", 12),
            ("数据日期", 14), ("创建时间", 20),
        ]

        for col_idx, (header_text, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 数据行
        keys = [
            "id", "title", "category", "priority", "sentiment",
            "view_count", "reply_count", "author_name", "source",
            "is_anomaly", "risk_level", "data_date", "created_at"
        ]

        for row_idx, post in enumerate(posts, 2):
            for col_idx, key in enumerate(keys, 1):
                value = post.get(key, "")
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                elif isinstance(value, bool):
                    value = "是" if value else "否"
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # 冻结首行
        ws.freeze_panes = "A2"

        # 自动筛选
        ws.auto_filter.ref = ws.dimensions

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
